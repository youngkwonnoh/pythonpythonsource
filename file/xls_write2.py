from openpyxl import Workbook

wb = Workbook()
print(wb.sheetnames)

# 시트 삭제
wb.remove(wb['Sheet'])

# 새로운 시트 생성
sheet1 = wb.create_sheet(title="매출전표")


wb.save("./data/test2.xlsx")