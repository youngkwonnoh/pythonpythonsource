import re
import openpyxl

# python VS java 라는 원본 문자열을 ' VS ' 를 기준으로 분리해보기
pattern = re.compile(" [A-Z]{2} ")
splited = pattern.split("python VS java")
print(splited)
print("\n")

# 801210-1011323 주민번호에서 - 기호를 * 로 바꿔서 출력하기
print(re.sub("-", "*", "801210-1011323"))
print("\n")

# data_kr.xlsx 의 첫번째 시트를 읽어서
# 주민번호 뒷자리를 * 로 바꿔서 출력하기
wb = openpyxl.load_workbook("./regex/data_kr.xlsx")
sheet1 = wb.active
pattern = re.compile("[0-9]{7}")

for row in sheet1.rows:
    # print(row[1].value)
    print(re.sub(pattern, "*******", row[1].value))

# train.xlsx를 읽어서 성별대로 4개의 시트를 만들어 정보 삽입하기
# train_gender.xlsx로 저장
# Mr(sheet 명은 남성), Miss(미혼여성), Mrs(기혼여성), Others(기타)

# 기존의 엑셀 파일 열기
wb = openpyxl.load_workbook("./regex/train.xlsx")
work_sheet = wb.active

# 새로운 엑셀 파일 작성
wb = openpyxl.Workbook()

# 기본 시트 활성화
wb_man = wb.active # 남자 승객 삽입
wb_man.title = '남성'
wb_man.column_dimensions['D'].width = 70

# 두번째 시트 생성
wb_solo_women = wb.create_sheet(title='미혼여성')
wb_solo_women.column_dimensions['D'].width = 70

# 세번째 시트 생성
wb_married_women = wb.create_sheet(title='기혼여성')
wb_married_women.column_dimensions['D'].width = 70

# 네번째 시트 생성
wb_others = wb.create_sheet(title='기타')
wb_others.column_dimensions['D'].width = 70

# 다섯번째 시트 생성
wb_report = wb.create_sheet(title='보고서')

pattern = re.compile(" [A-Za-z]+\.")

list1 = []

# 생존자 수, 사망자 수
man_survived, man_unsurvived = 0, 0
single_women_survived, single_women_unsurvived = 0, 0
married_women_survived, married_women_unsurvived = 0, 0
others_survived, others_unsurvived = 0, 0



for row in work_sheet.rows:
    # print(row[0], row[0].row, row[0].coordinate, row[0].value)

    # 정규식과 매치되는 문자열 추출
    data = pattern.findall(row[3].value)

    #제목행 복사 후 각 시트에 붙여넣기
    if row[0].row == 1:
        for title in row:
            # print(title.value, end="\t")
            list1.append(title.value)
        # print(list1)
        wb_man.append(list1)
        wb_solo_women.append(list1)
        wb_married_women.append(list1)
        wb_others.append(list1)
    else:
        if data[0] == ' Mr.':
            wb_man.append([item.value for item in row])
            if row[1].value == 1:
                man_survived += 1
            else:
                man_unsurvived += 1
        elif data[0] == ' Miss.':
            wb_solo_women.append([item.value for item in row])
            if row[1].value == 1:
                single_women_survived += 1
            else:
                single_women_unsurvived += 1
        elif data[0] == ' Mrs.':
            wb_married_women.append([item.value for item in row])
            if row[1].value == 1:
                married_women_survived += 1
            else:
                married_women_unsurvived += 1
        else:
            wb_others.append([item.value for item in row])
            if row[1].value == 1:
                others_survived += 1
            else:
                others_unsurvived += 1

# 보고서 시트 작성
wb_report.append(['분류', '생존자수', '사망자수', '생존률'])

# 남성 생존률 계산 ex) 25.35%
survived_rate = "%.2f%%" % (man_survived / (man_survived + man_unsurvived) * 100)
wb_report.append(['남성', man_survived,  man_unsurvived, survived_rate])

# 미혼여성 생존률 계산 ex) 25.35%
survived_rate = "%.2f%%" % (single_women_survived / (single_women_survived +single_women_unsurvived) * 100)
wb_report.append(['미혼여성', single_women_survived, single_women_unsurvived, survived_rate])

# 기혼여성 생존률 계산 ex) 25.35%
survived_rate = "%.2f%%" % (married_women_survived / (married_women_survived + married_women_unsurvived) * 100)
wb_report.append(['미혼여성', married_women_survived,  married_women_unsurvived, survived_rate])

# 기타 생존률 계산 ex) 25.35%
survived_rate = "%.2f%%" % (others_survived / (others_survived + others_unsurvived) * 100)
wb_report.append(['기타', others_survived,  others_unsurvived, survived_rate])

wb.save("./regex/train_gender.xlsx")